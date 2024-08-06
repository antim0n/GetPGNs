# import openpyxl module
import openpyxl
 
# Give the location of the file
path = input("Excel File Name:")
filename = "save.pgn"
# starting at (2;6) in an excel file the whole column is saved as long as there is data
# e.g. column 6: -> e4 \ e5 \ Nf3 \ ..., other formats possible
excel_row = 2
excel_column = 6
 
# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
 
# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
 
cell_obj = sheet_obj.cell(row=excel_row, column=excel_column)

# save in file
while cell_obj.value is not None:
    with open(filename, "a") as f:
        f.write(str(cell_obj.value) + "\r\n")
    excel_row += 1
    cell_obj = sheet_obj.cell(row=excel_row, column=excel_column)
 

#open and read the file after the appending:
#f = open("save.pgn", "r")
#print(f.read())